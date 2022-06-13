# Importing required libraries
from email import header
from importlib.resources import path
from openpyxl import Workbook
import pandas as pd
import openpyxl as xl
from pathlib import Path

# Function to get summary of the Excel file
def find_Min_Max_Mn(sheetname):
    sheet= Wb[sheetname]
    sheet['G2']= "Min_Temp"
    sheet['G3']= "Max_Temp"
    sheet['G4']= "Mean_Temp"
    sheet['G6']= "Min_Date"
    sheet['G7']= "Max_Date"
    sheet['H2']= "=MIN(C:C)"
    sheet['H3']= "=MAX(C:C)"
    sheet['H4']= "=AVERAGE(C:C)"
    sheet['H6']= "=MIN(A:A)"
    sheet['H7']= "=MAX(A:A)"

# Assigning a name to excel file    
Filename = "BCM.xlsx"

# Creating a path
mycsvdir = str(Path.cwd())
#print (mycsvdir)
Filepath = mycsvdir+  '/output/' + Filename
#print (Filepath)
csvfiles= Path(mycsvdir+'/data/logs').glob('*.csv')
Wb=xl.Workbook()
Wb.save(filename=Filepath)
excelfile= pd.ExcelWriter(Filepath)

# Importing csv file to excel file in a different sheet
for csvfile in csvfiles:
    sheetname=csvfile.stem
    df=pd.read_csv(csvfile, header=None, names=["datetine","scale","temaratere"])
    df.to_excel(excelfile,sheet_name=sheetname,index=None)
excelfile.close()

# Finding Min and Max of the variables Temperature and Date
Wb=xl.load_workbook(Filepath)
for i in range (len(Wb.sheetnames)):
    find_Min_Max_Mn(Wb.sheetnames[i])
Wb.save(filename=Filepath)



    



