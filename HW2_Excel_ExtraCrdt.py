# Importing required libraries
import pandas as pd
import openpyxl as xl
from pathlib import Path

# Creating a path
mycsvdir = str(Path.cwd())
print (mycsvdir)
Filepath = mycsvdir+  '/output/'
print (Filepath)
csvfiles= Path(mycsvdir+'/data/extracredit_Data').glob('*.csv')
Filenames=[]

# Function to get summary of the Excel file
def find_Min_Max_Mn(Wb,sheetname):
    sheet= Wb[sheetname]
    sheet['G2']= "Min_Temp"
    sheet['G3']= "Max_Temp"
    sheet['G4']= "Mean_Temp"
    sheet['G6']= "Min_Date"
    sheet['G7']= "Max_Date"
    sheet['H2']= "=MIN(C:C)"
    sheet['H3']= "=MAX(C:C)"
    sheet['H4']= "=AVERAGE(C:C)"
    sheet['H6']= "=MIN(A:A)"
    sheet['H7']= "=MAX(A:A)"

# Function for Splitting the csv filename
def Get_streams():
    for csvfile in csvfiles:
        sheet=csvfile.stem  
    # print(sheetname)
        Filestream = sheet.split("-")
        if Filestream[0] not in Filenames:
            Filenames.append(Filestream[0])

# Function for reading csv file saving it as an excel file 
def importWS(i,excelfile,Filepath):
    csvfiles= Path(mycsvdir+'/data/extracredit_Data').glob('*.csv')
    for csvfile in csvfiles:
        sheetname=csvfile.stem
        name=sheetname.split("-")
        if Filenames[i] == name[0]:
            df=pd.read_csv(csvfile, header=None, names=["datetine","scale","temaratere"])
            df.to_excel(excelfile,sheet_name=sheetname,index=None)
    excelfile.close()
    Wb=xl.load_workbook(Filepath)
    for i in range (len(Wb.sheetnames)):
        find_Min_Max_Mn(Wb,Wb.sheetnames[i])
    Wb.save(filename=Filepath)
            
# Function for creating excel file 
def CreatXL():
    for i in range (len(Filenames)): 
        Filepath = mycsvdir+  '/output/' + Filenames[i] +".xlsx"
        #print (Filepath)
        Wb=xl.Workbook()
        Wb.save(filename=Filepath)
        excelfile= pd.ExcelWriter(Filepath)
        importWS(i,excelfile,Filepath)
    
    
Get_streams()
print(Filenames)
print(len(Filenames))
CreatXL()